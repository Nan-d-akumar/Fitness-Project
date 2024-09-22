from django.contrib.auth.models import User
from django.http import HttpResponse
from django.shortcuts import render, redirect
from openpyxl.workbook import Workbook

from adminapp.models import Tbl_staff
from homeapp.models import Tbl_Category, Tbl_customer, Tbl_Trainer, Tbl_SubCategory,Tbl_tutorial,Tbl_diet,Tbl_payment,Tbl_card,Tbl_Feedback,Tbl_trainer_assign
from django.http import HttpResponse
from django.template.loader import get_template
from django.conf import settings
from xhtml2pdf import pisa
from io import BytesIO
from django.contrib import messages



# Create your views here.

def adminhome(request,tot=0):
    s = len(Tbl_staff.objects.all())
    c = len(Tbl_customer.objects.all())
    cat = len(Tbl_Trainer.objects.filter(trainer_approval=True))
    tu = len(Tbl_tutorial.objects.all())
    tp = Tbl_payment.objects.all()
    for i in tp:
        tot += i.amount

    return render(request, 'admindash.html',{'s':s,'c':c,'cat':cat,'tu':tu,'pyt':tp,'tot':tot})


def categoryall(request):
    call = Tbl_Category.objects.all()
    return render(request, 'categoryall.html', {'c': call})


# Category add
def categoryadd(request):
    if request.method == 'POST':
        cn = request.POST['cname']
        cim = request.FILES['cimg']
        cd = request.POST['cdesc']
        if Tbl_Category.objects.filter(category_name=cn).exists():
            print('Category already exists')
        else:
            c = Tbl_Category(category_name=cn, category_image=cim, category_desc=cd)
            c.save()
            return redirect('catall')
    return render(request, 'categoryadd.html')


# Customer Show

def customerall(request):
    call = Tbl_customer.objects.all()
    return render(request, 'customerall.html', {'cu': call})


# Customer edit

def customeredit(request, cid):
    cus = Tbl_customer.objects.get(id=cid)
    if request.method == 'POST':
        cust_fname = request.POST['cust_fname']
        cust_lname = request.POST['cust_lname']
        cust_age = request.POST['cust_age']
        cust_height = request.POST['cust_height']
        cust_weight = request.POST['cust_weight']
        cust_phno = request.POST['cust_phno']
        cust_prob = request.POST['cust_prob']
        cust_email = request.POST['cust_email']
        cust_gender = request.POST['gender']
        cus.cust_fname = cust_fname
        cus.cust_lname = cust_lname
        cus.cust_age = cust_age
        cus.cust_height = cust_height
        cus.cust_weight = cust_weight
        cus.cust_gender = cust_gender
        cus.cust_phone = cust_phno
        cus.cust_problem = cust_prob
        cus.cust_email = cust_email
        cus.save()
        print('Succesfully updated')
        if request.user.is_superuser:
            return redirect('custall')
        else:
            return redirect('/')

    return render(request, 'custedit.html', {'c': cus})


# customer deactivate

def CustDeactive(request, did):
    cus = Tbl_customer.objects.get(id=did)
    cus.cust_status = False
    cus.save()
    print('deactivate')
    return redirect('custall')


# Customer activate

def CustActivate(request, did):
    cus = Tbl_customer.objects.get(id=did)
    cus.cust_status = True
    cus.save()
    return redirect('custall')


# Trainer Approved List

def Trainer_Approved_List(request):
    trainer = Tbl_Trainer.objects.filter(trainer_approval=True)
    return render(request, 'trainerall.html', {'tr': trainer})


# Approvals List

def Trainer_Approvals(request):
    call = Tbl_Trainer.objects.filter(trainer_approval=False)
    return render(request, 'trainer_approval.html', {'tr': call})


# Trainer Approving Function

def Trainer_Approving_function(request, aid):
    data = Tbl_Trainer.objects.get(id=aid)
    data.trainer_approval = True
    data.trainer_status = True
    data.save()
    return redirect('ta')


# staff registration function
def staff_registration(request):
    if request.method == 'POST':
        staff_fname = request.POST['staff_fname']
        staff_lname = request.POST['staff_lname']
        staff_age = request.POST['staff_age']
        staff_phno = request.POST['staff_phno']
        staff_email = request.POST['staff_email']
        staff_pass = request.POST['staff_pass']
        staff_repass = request.POST['staff_repass']
        staff_gender = request.POST['gender']
        if User.objects.filter(username=staff_email).exists():
            print('Email Already exist')
        else:
            if staff_pass == staff_repass:
                user = User.objects.create_user(first_name=staff_fname, last_name=staff_lname, username=staff_email,
                                                password=staff_pass,is_staff=True)
                user.save()
                register = Tbl_staff(staff_fname=staff_fname, staff_lname=staff_lname, staff_age=staff_age,
                                     staff_phone=staff_phno,
                                     staff_email=staff_email, user=user, staff_gender=staff_gender, staff_status=True)
                register.save()
                return redirect('staffall')
            else:
                print('password not matching')
                return redirect('staff_reg')

    return render(request, 'staff_registration.html')


# Staff show
def staffall(request):
    call = Tbl_staff.objects.all()
    return render(request, 'staffall.html', {'cu': call})


# Staff edit
def staffedit(request, cid):
    cus = Tbl_staff.objects.get(id=cid)
    if request.method == 'POST':
        staff_fname = request.POST['staff_fname']
        staff_lname = request.POST['staff_lname']
        staff_age = request.POST['staff_age']
        staff_phno = request.POST['staff_phno']
        staff_email = request.POST['staff_email']
        staff_gender = request.POST['gender']
        cus.staff_fname = staff_fname
        cus.staff_lname = staff_lname
        cus.staff_age = staff_age
        cus.staff_gender = staff_gender
        cus.staff_phone = staff_phno
        cus.staff_email = staff_email
        cus.save()
        print('Succesfully updated')
        if request.user.is_superuser:
            return redirect('staffall')
        else:
            return redirect('/')

    return render(request, 'staffedit.html', {'c': cus})


# staff deactive
def staffDeactive(request, did):
    cus = Tbl_staff.objects.get(id=did)
    cus.staff_status = False
    cus.save()
    print('deactivate')
    return redirect('staffall')


# Staff activate

def staffActivate(request, did):
    cus = Tbl_staff.objects.get(id=did)
    cus.staff_status = True
    cus.save()
    return redirect('staffall')


def subcategoryadd(request):
    cat = Tbl_Category.objects.all()
    if request.method == 'POST':
        sname = request.POST['subname']
        category = request.POST['category']
        cato = Tbl_Category.objects.get(id=category)
        sc = Tbl_SubCategory(category=cato, sub_category_name=sname)
        sc.save()
        return redirect('subcatall')
    return render(request, 'subcategoryadd.html', {'cat': cat})


# sub category all
def subcategoryall(request):
    call = Tbl_SubCategory.objects.all()
    return render(request, 'subcatall.html', {'c': call})

#subcategory edit
def subcatedit(request, cid):
    cus = Tbl_SubCategory.objects.get(id=cid)
    cat = Tbl_Category.objects.all()
    if request.method == 'POST':
        subcategory_name = request.POST['subname']
        category_name = request.POST['cname']
        cus.category=Tbl_Category.objects.get(id=category_name)
        cus.sub_category_name=subcategory_name
        cus.save()
        print('Succesfully updated')
        return redirect('subcatall')
    return render(request, 'subcatedit.html', {'cus': cus,'cat':cat})


# category edit
def catedit(request, cid):
    cus = Tbl_Category.objects.get(id=cid)
    if request.method == 'POST':
        category_name = request.POST['cname']
        category_desc = request.POST['cdesc']
        img = request.FILES['cimage']
        cus.category_name=category_name
        cus.category_desc=category_desc
        cus.category_image=img
        cus.save()
        print('Succesfully updated')
        return redirect('catall')
    return render(request, 'categoryedit.html', {'cus': cus})

#add tutorials
def tutorialadd(request):
    scat = Tbl_SubCategory.objects.all()
    if request.method == 'POST':
        sub_category_name = request.POST['subcat']
        exercise_name = request.POST['exercise_name']
        video = request.POST['video']
        reps = request.POST['reps']
        sets = request.POST['sets']
        rest_time = request.POST['rest_time']
        sub = Tbl_SubCategory.objects.get(id=sub_category_name)
        tu = Tbl_tutorial(sub_category_name=sub,video=video,exercise_name=exercise_name,
                             reps=reps,sets=sets,rest_time=rest_time)
        tu.save()
        return redirect('tall')
    return render(request, 'tutorialadd.html', {'scat': scat})

#tutorial all
def tutorialall(request):
    call = Tbl_tutorial.objects.all()
    return render(request, 'tutorialall.html', {'c': call})

#tutorial edit
def tutorialedit(request, id):
    cus = Tbl_tutorial.objects.get(id=id)
    s = Tbl_SubCategory.objects.all()
    if request.method == 'POST':
        sub_category_name = request.POST['subcat']
        exercise_name = request.POST['exercise_name']
        video = request.FILES['video']
        reps = request.POST['reps']
        sets = request.POST['sets']
        rest_time = request.POST['rest_time']
        cus.sub_category_name=Tbl_SubCategory.objects.get(id=sub_category_name)
        cus.exercise_name = exercise_name
        cus.video = video
        cus.reps = reps
        cus.sets = sets
        cus.rest_time = rest_time
        cus.save()
        print('Succesfully updated')
        return redirect('tall')
    return render(request, 'tutorialedit.html', {'scat': cus,'s':s})


#customer exel
def database_to_excelcust(request):
    # Fetch data from the Customer model
    queryset = Tbl_customer.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = ['cust_fname', 'cust_age', 'cust_email', 'cust_phone', 'cust_gender']
    ws.append(headers)

    # Write data to the worksheet
    for customer in queryset:
        data = []
        for header in headers:
            # Handle the ForeignKey field (user) by extracting the username
            if header == 'user':
                user_value = getattr(customer, header)
                if user_value:
                    data.append(user_value.username)
                else:
                    data.append(None)
            else:
                data.append(getattr(customer, header))

        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_data.xlsx'
    wb.save(response)

    return response




def database_to_excelstaff(request):
    # Fetch data from the Customer model
    queryset = Tbl_staff.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = ['staff_fname', 'staff_age', 'staff_gender', 'staff_phone', 'staff_email']
    ws.append(headers)

    # Write data to the worksheet
    for customer in queryset:
        data = []
        for header in headers:
            # Handle the ForeignKey field (user) by extracting the username
            if header == 'user':
                user_value = getattr(customer, header)
                if user_value:
                    data.append(user_value.username)
                else:
                    data.append(None)
            else:
                data.append(getattr(customer, header))

        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=staff_data.xlsx'
    wb.save(response)

    return response


def database_to_exceltrainer(request):
    # Fetch data from the Customer model
    queryset = Tbl_Trainer.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = ['trainer_name', 'trainer_age', 'trainer_level', 'trainer_email', 'trainer_phone']
    ws.append(headers)

    # Write data to the worksheet
    for customer in queryset:
        data = []
        for header in headers:
            # Handle the ForeignKey field (user) by extracting the username
            if header == 'user':
                user_value = getattr(customer, header)
                if user_value:
                    data.append(user_value.username)
                else:
                    data.append(None)
            else:
                data.append(getattr(customer, header))

        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=trainer_data.xlsx'
    wb.save(response)

    return response

def database_to_excelsales(request):
    # Fetch data from the Tbl_payment model
    queryset = Tbl_payment.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = ['customer', 'type', 'date', 'amount']
    ws.append(headers)

    # Write data to the worksheet
    for payment in queryset:
        data = []
        # Access the related customer and get the cust_fname
        if payment.card and payment.card.customer:
            customer_name = payment.card.customer.first_name
            data.append(customer_name)
        else:
            data.append('N/A')  # Provide a default value if customer data is missing
        data.append(payment.type)
        data.append(payment.date)
        data.append(payment.amount)
        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=trainer_data.xlsx'
    wb.save(response)

    return response
def database_to_excelcat(request):
    # Fetch data from the Customer model
    queryset = Tbl_Category.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = ['category_name', 'category_desc']
    ws.append(headers)

    # Write data to the worksheet
    for customer in queryset:
        data = []
        for header in headers:
            # Handle the ForeignKey field (user) by extracting the username
            if header == 'user':
                user_value = getattr(customer, header)
                if user_value:
                    data.append(user_value.username)
                else:
                    data.append(None)
            else:
                data.append(getattr(customer, header))

        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=category_data.xlsx'
    wb.save(response)

    return response

def database_to_excelsubcat(request):
    # Fetch data from the Customer model
    queryset = Tbl_SubCategory.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = ['category', 'sub_category_name']
    ws.append(headers)

    # Write data to the worksheet
    for customer in queryset:
        data = []
        for header in headers:
            # Handle the ForeignKey field (user) by extracting the username
            if header == 'category':
                user_value = getattr(customer, header)
                if user_value:
                    data.append(user_value.category_name)
                else:
                    data.append(None)
            else:
                data.append(getattr(customer, header))

        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=subcategory_data.xlsx'
    wb.save(response)

    return response

def database_to_exceltutorial(request):
    # Fetch data from the Customer model
    queryset = Tbl_tutorial.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = ['sub_category_name', 'exercise_name','reps','sets','rest_time']
    ws.append(headers)

    # Write data to the worksheet
    for customer in queryset:
        data = []
        for header in headers:
            # Handle the ForeignKey field (user) by extracting the username
            if header == 'tutorial':
                user_value = getattr(customer, header)
                if user_value:
                    data.append(user_value.exercise_name)
                else:
                    data.append(None)
            else:
                data.append(getattr(customer, header))

        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=tutorial_data.xlsx'
    wb.save(response)

    return response
# Staff diet function

def StaffDietRequest(request):
    req = Tbl_diet.objects.filter(status=False)
    return render(request,'staffDiet.html',{'data':req})


# staff request confirmation

def StaffRequestConfirmation(request,id):

    data = Tbl_diet.objects.get(id=id)
    c = Tbl_customer.objects.get(user=data.customer)
    if request.method == 'POST':
        dietfile = request.FILES.get('doc',None)
        if dietfile == None:
            data.status=False
            data.save()
            return redirect('streq')
        else:
            data.diet=dietfile
            data.status=True
            data.staff = request.user
            data.save()
            return redirect('streq')

    return render(request,'staffDiet2.html',{'d':data,'c':c})


def generate_pdf(request):
    cust = Tbl_customer.objects.all()
    # Get data for your template, you can replace it with your own dynamic data retrieval logic
    context = {
        'cust': cust
    }

    # Render the template
    template = get_template('customerpdf.html')
    html_content = template.render(context)

    # Create a PDF file
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'filename="your_pdf_filename.pdf"'
        return response

    return HttpResponse('Error generating PDF: %s' % pdf.err)


def generate_pdfstaff(request):
    cust = Tbl_staff.objects.all()
    # Get data for your template, you can replace it with your own dynamic data retrieval logic
    context = {
        'cust': cust
    }

    # Render the template
    template = get_template('staffpdf.html')
    html_content = template.render(context)

    # Create a PDF file
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'filename="your_pdf_filename.pdf"'
        return response

    return HttpResponse('Error generating PDF: %s' % pdf.err)


def generate_pdftrainer(request):
    cust = Tbl_Trainer.objects.all()
    # Get data for your template, you can replace it with your own dynamic data retrieval logic
    context = {
        'cust': cust
    }

    # Render the template
    template = get_template('trainerpdf.html')
    html_content = template.render(context)

    # Create a PDF file
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'filename="your_pdf_filename.pdf"'
        return response

    return HttpResponse('Error generating PDF: %s' % pdf.err)

def generate_pdfcategory(request):
    cust = Tbl_Category.objects.all()
    # Get data for your template, you can replace it with your own dynamic data retrieval logic
    context = {
        'cust': cust
    }

    # Render the template
    template = get_template('categorypdf.html')
    html_content = template.render(context)

    # Create a PDF file
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'filename="your_pdf_filename.pdf"'
        return response

    return HttpResponse('Error generating PDF: %s' % pdf.err)

def generate_pdfsubcategory(request):
    cust = Tbl_SubCategory.objects.all()
    # Get data for your template, you can replace it with your own dynamic data retrieval logic
    context = {
        'cust': cust
    }

    # Render the template
    template = get_template('subcatpdf.html')
    html_content = template.render(context)

    # Create a PDF file
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'filename="your_pdf_filename.pdf"'
        return response

    return HttpResponse('Error generating PDF: %s' % pdf.err)

def generate_pdftutorial(request):
    cust = Tbl_tutorial.objects.all()
    # Get data for your template, you can replace it with your own dynamic data retrieval logic
    context = {
        'cust': cust
    }

    # Render the template
    template = get_template('tutorialpdf.html')
    html_content = template.render(context)

    # Create a PDF file
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'filename="your_pdf_filename.pdf"'
        return response

    return HttpResponse('Error generating PDF: %s' % pdf.err)

def generate_pdfsales(request):
    cust = Tbl_payment.objects.all()
    # Get data for your template, you can replace it with your own dynamic data retrieval logic
    context = {
        'cust': cust
    }

    # Render the template
    template = get_template('salespdf.html')
    html_content = template.render(context)

    # Create a PDF file
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'filename="your_pdf_filename.pdf"'
        return response

    return HttpResponse('Error generating PDF: %s' % pdf.err)


def generate_pdfbilldiet(request,id):
    cust = Tbl_payment.objects.get(id=id)
    # Get data for your template, you can replace it with your own dynamic data retrieval logic
    context = {
        'cust': cust
    }

    # Render the template
    template = get_template('Billdiet2.html')
    html_content = template.render(context)

    # Create a PDF file
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'filename="your_pdf_filename.pdf"'
        return response

    return HttpResponse('Error generating PDF: %s' % pdf.err)

def generate_pdfbilltrainer(request,id):
    cust = Tbl_payment.objects.get(id=id)
    ta = Tbl_trainer_assign.objects.all()
    # Get data for your template, you can replace it with your own dynamic data retrieval logic
    context = {
        'cust': cust,
        'ta':ta
    }

    # Render the template
    template = get_template('billtrainer2.html')
    html_content = template.render(context)

    # Create a PDF file
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'filename="your_pdf_filename.pdf"'
        return response

    return HttpResponse('Error generating PDF: %s' % pdf.err)

def feedbackview(request):
    call = Tbl_Feedback.objects.all()
    return render(request, 'feedbackview.html',{'call':call})
